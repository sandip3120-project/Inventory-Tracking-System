from django.contrib import admin
from django.urls import reverse
from django.http import HttpResponse
from io import BytesIO
import qrcode
from .models import SiteConfig, ReconciliationLog

from .models import Material, Batch, Customer, Roll, Location, Transaction, Department, Profile
import io
from django.urls import path
from openpyxl import Workbook



@admin.register(Material)
class MaterialAdmin(admin.ModelAdmin):
#    list_display  = ('material_number', 'description', 'colour')
#    search_fields = ('material_number', 'description')
    list_display  = ('material_number', 'description',
                      'department', 'created_by', 'created_at')
    readonly_fields = ('created_by','created_at')
    search_fields = ('material_number', 'description')

    def save_model(self, request, obj, form, change):
        # if someone creates a Material by hand in the admin,
        # also set created_by / created_at
        if not change or not obj.created_by:
            obj.created_by = request.user
        super().save_model(request, obj, form, change)

@admin.register(Batch)
class BatchAdmin(admin.ModelAdmin):
    list_display  = ('batch_number', 'material', 'created_at')
    search_fields = ('batch_number',)
    list_filter   = ('material',)


@admin.register(Customer)
class CustomerAdmin(admin.ModelAdmin):
    list_display  = ('name',)
    search_fields = ('name',)


@admin.register(Roll)
class RollAdmin(admin.ModelAdmin):
    list_display  = ('roll_id', 'batch', 'weight_kg', 'current_location', 'status')
    search_fields = ('roll_id', 'batch__batch_number')
    list_filter   = ('status', 'current_location')


def download_location_qr(modeladmin, request, queryset):
    """
    Admin action: generate a PNG QR for the first selected Location.
    """
    loc = queryset.first()
    if not loc:
        return
    # Build the absolute URL for scanning this location
    url = request.build_absolute_uri(
        reverse('location-scan', args=[loc.location_code])
    )
    # Generate QR image
    img = qrcode.make(url)
    buf = BytesIO()
    img.save(buf, format='PNG')
    buf.seek(0)

    response = HttpResponse(buf.read(), content_type='image/png')
    response['Content-Disposition'] = f'attachment; filename="{loc.location_code}.png"'
    return response

download_location_qr.short_description = "Download QR for selected Location"

@admin.register(Location)
class LocationAdmin(admin.ModelAdmin):
    list_display  = (
        'location_code',
        'get_dept_code',   # show the two‑letter code
        'row',
        'column',
        'type',
    )
    search_fields = ('location_code',)
    list_filter   = ('type', 'department__code')  # you can also filter by dept
    actions       = [download_location_qr]

    def get_dept_code(self, obj):
        return obj.department.code
    get_dept_code.short_description = 'Dept'
    get_dept_code.admin_order_field = 'department__code'



@admin.register(Transaction)
class TransactionAdmin(admin.ModelAdmin):
    list_display  = ('roll', 'action', 'location', 'user', 'scanned_at')
    search_fields = ('roll__roll_id', 'user', 'action')
    list_filter   = ('action',)


# Department
@admin.register(Department)
class DepartmentAdmin(admin.ModelAdmin):
    list_display = ('name','code')



# warehouse/admin.py

from django.contrib import admin
from django.contrib.auth.models import Group
from .models import Profile, Material, Batch, Customer, Roll, Location, Transaction, Department

# … your existing admin registrations …

@admin.register(Profile)
class ProfileAdmin(admin.ModelAdmin):
    list_display  = ('user','requested_group','department','extra_access','needs_approval')
    list_filter   = ('requested_group','department','needs_approval')
    search_fields = ('user__username','user__email')
    fieldsets = (
        (None, {
            'fields': (
                'user','phone','pin','id_card','needs_approval','requested_group'
            )
        }),
        ('Department Assignment', {
            'fields': ('department','extra_access'),
            'description': "Give each Plant Manager/Operator/Stock Keeper their home dept and any extra_access."
        }),
    )

    def save_model(self, request, obj, form, change):
        # First save the Profile itself
        print(f"[DEBUG] About to save Profile: user={obj.user.username!r}, "
              f"requested_group={obj.requested_group!r}, department={obj.department!r}")
        super().save_model(request, obj, form, change)
        fresh = Profile.objects.get(pk=obj.pk)
        print(f"[DEBUG] Saved Profile: user={fresh.user.username!r}, "
              f"department={fresh.department!r}")
        # Now sync the User’s groups to match requested_group
        ITS_GROUPS = [
            'Factory Admin',
            'Plant Manager',
            'Stock Keeper',
            'Operator',
            'Forklift Driver',
            'Dept SK',
            'View Only',
        ]

        # Remove the user from any of those groups
        obj.user.groups.remove(*Group.objects.filter(name__in=ITS_GROUPS))

        # If they picked one, add them into that group
        if obj.requested_group:
            try:
                g = Group.objects.get(name=obj.requested_group)
                obj.user.groups.add(g)
            except Group.DoesNotExist:
                self.message_user(
                    request,
                    f"⚠️ Group “{obj.requested_group}” doesn’t exist in Django’s auth.Group!",
                    level='error'
                )


@admin.register(ReconciliationLog)
class ReconciliationLogAdmin(admin.ModelAdmin):
    list_display = ('run_at','is_clean')
    readonly_fields = ('run_at','is_clean','mismatches')



# ───  MASTER AUDIT EXPORT VIEW  ───────────────────────────────────────────────

def master_export(request):
    """
    In‑memory Excel export:
      • Sheet1: Materials + roll metadata
      • Sheet2: Full Transactions log
      • Sheet3: Summary aggregates
    """
    # 1) workbook + first sheet (Materials)
    wb  = Workbook()
    ws1 = wb.active
    ws1.title = "Materials"
    ws1.append([
        "Roll ID", "Material #", "Description",
        
        "Batch #", "Weight (kg)", "Current Location",
        "Status", "Posting Date", "Dispatch Customer"
    ])
    # defer models import until runtime
    from .models import Roll, Transaction
    for r in Roll.objects.select_related('batch__material'):
        last_tx = r.transaction_set.order_by('-scanned_at').first()
        posting_date   = last_tx.scanned_at if last_tx else None
        dispatch_cust  = (last_tx.customer.name if last_tx and last_tx.action=="DISPATCH" else "")
        ws1.append([
            str(r.roll_id),
            r.batch.material.material_number,
            r.batch.material.description,
            r.batch.batch_number,
            r.weight_kg,
            r.current_location or "",
            r.status,
            posting_date.strftime("%Y-%m-%d %H:%M") if posting_date else "",
            dispatch_cust,
        ])

    # 2) Transactions sheet
    ws2 = wb.create_sheet("Transactions")
    ws2.append([
        "TX ID", "Roll ID", "Action",
        "Location Code", "Customer",
        "User", "Scanned At"
    ])
    for tx in Transaction.objects.select_related('roll','location','customer').order_by('scanned_at'):
        ws2.append([
            tx.id,
            str(tx.roll.roll_id),
            tx.action,
            tx.location.location_code if tx.location else "",
            tx.customer.name if tx.customer else "",
            tx.user,
            tx.scanned_at.strftime("%Y-%m-%d %H:%M"),
        ])

    # 3) Summary sheet
    ws3 = wb.create_sheet("Summary")
    ws3.append(["Metric","Count"])
    total_rolls = Roll.objects.count()
    total_tx    = Transaction.objects.count()
    ws3.append(["Total Rolls", total_rolls])
    ws3.append(["Total Transactions", total_tx])
    ws3.append([])
    ws3.append(["Transactions by Action","Count"])
    from django.db.models import Count
    for row in Transaction.objects.values('action').annotate(c=Count('id')):
        ws3.append([row['action'], row['c']])

    # 4) stream it out
    stream = io.BytesIO()
    wb.save(stream)
    response = HttpResponse(
        stream.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="ITS_master_audit.xlsx"'
    return response


# ───  INJECT EXPORT URL INTO DEFAULT ADMIN  ────────────────────────────────────

# grab the original get_urls
_original_admin_urls = admin.site.get_urls

def get_admin_urls():
    """
    Prepend our export URL so it lives at:
      /admin/warehouse/master_export/
    """
    custom = [
        path(
            'warehouse/master_export/',
            admin.site.admin_view(master_export),
            name='warehouse_master_export'
        )
    ]
    return custom + _original_admin_urls()

# override the admin site's URL resolver
admin.site.get_urls = get_admin_urls
